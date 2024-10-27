import akshare as ak
import pandas as pd
import datetime
from openpyxl import load_workbook

# 应用样式，涨为红，跌为绿
def color_negative_green_positive_red(value):
    if float(value.strip('%')) > 0:
        return 'color: red'
    elif float(value.strip('%')) < 0:
        return 'color: green'
    else:
        return ''

# 获取热门股票排名数据
stock_hot_rank_em_df = ak.stock_hot_rank_em()

# 将代码列转换为字符串类型，并去掉股票代码前面的SZ或SH
stock_hot_rank_em_df['代码'] = stock_hot_rank_em_df['代码'].astype(str).str[2:]

# 删除不需要的列
stock_hot_rank_em_df.drop(columns=['涨跌额'], inplace=True)

# 将涨跌幅从字符串转换为数值
stock_hot_rank_em_df['涨跌幅'] = stock_hot_rank_em_df['涨跌幅'].replace('%', '').astype(float)

# 保存到 Excel 文件
stock_hot_rank_em_df.to_excel(r'C:\Users\65179\Desktop\quant\每日热股排名.xlsx', index=False)

# 读取 Excel 文件
hot_stocks_df = pd.read_excel(r'C:\Users\65179\Desktop\quant\每日热股排名.xlsx', dtype=str)
company_info_df = pd.read_excel(r'C:\Users\65179\Desktop\quant\代码行业概念.xlsx', dtype=str)

# 根据代码列进行合并
merged_df = pd.merge(hot_stocks_df, company_info_df[['代码', '行业', '所属概念']], on='代码', how='left')


business_info_df = pd.read_excel(r'C:\Users\65179\Desktop\quant\代码主营业务.xlsx', dtype=str)


merged_df = pd.merge(merged_df, business_info_df[['股票代码','主营业务','产品类型', '产品名称', '经营范围']], left_on='代码', right_on='股票代码', how='left')
merged_df.drop(columns=['股票代码'], inplace=True)
merged_df['涨跌幅'].fillna(0, inplace=True)
merged_df['涨跌幅'] = merged_df['涨跌幅'].astype(str) + '%'


merged_df = merged_df.style
merged_df = merged_df.map(color_negative_green_positive_red, subset=['涨跌幅'])




# 保存合并后的数据到新的 Excel 文件
output_file_merged = r'C:\Users\65179\Desktop\quant\每日热股排名.xlsx'
merged_df.to_excel(output_file_merged, sheet_name="热股排名", index=False)
print(f"合并后的每日热股排名数据已成功保存到 {output_file_merged}")

# 当日连板天梯获取
date = datetime.date.today().strftime("%Y%m%d")
stock_zt_pool_em_df = ak.stock_zt_pool_em(date=date)

# 数据处理

stock_zt_pool_em_df['成交额'] = (stock_zt_pool_em_df['成交额'] / 100000000).round(2).astype(str) + '亿'
stock_zt_pool_em_df['流通市值'] = (stock_zt_pool_em_df['流通市值'] / 100000000).round(2).astype(str) + '亿'
stock_zt_pool_em_df['总市值'] = (stock_zt_pool_em_df['总市值'] / 100000000).round(2).astype(str) + '亿'
stock_zt_pool_em_df['封板资金'] = (stock_zt_pool_em_df['封板资金'] / 100000000).round(2).astype(str) + '亿'
stock_zt_pool_em_df['涨跌幅'] = stock_zt_pool_em_df['涨跌幅'].round(2).astype(str) + '%'
stock_zt_pool_em_df['换手率'] = stock_zt_pool_em_df['换手率'].round(2).astype(str) + '%'
stock_zt_pool_em_df['首次封板时间'] = pd.to_datetime(stock_zt_pool_em_df['首次封板时间'], format='%H%M%S').dt.time
stock_zt_pool_em_df['最后封板时间'] = pd.to_datetime(stock_zt_pool_em_df['最后封板时间'], format='%H%M%S').dt.time
stock_zt_pool_em_df['连板数'] = stock_zt_pool_em_df['连板数'].astype(int)
stock_zt_pool_em_df.sort_values(by=['连板数', '最后封板时间'], ascending=[False, True], inplace=True)
max_value = stock_zt_pool_em_df['序号'].max()
max_lianban = stock_zt_pool_em_df['连板数'].max()
stock_zt_pool_em_df.drop(columns=['序号'], inplace=True)




# 改变涨跌幅颜色

stock_zt_pool_em_df.style
stock_zt_pool_em_df = stock_zt_pool_em_df.style.map(color_negative_green_positive_red, subset=['涨跌幅'])

# 加载已存在的Excel文件
book = load_workbook(output_file_merged)

# 创建一个新的工作表
ws = book.create_sheet("连板天梯")

# 将DataFrame写入新工作表
with pd.ExcelWriter(output_file_merged, mode='a', engine='openpyxl') as writer:
    stock_zt_pool_em_df.to_excel(writer, sheet_name='连板天梯', index=False)
    print(f"连板天梯数据已成功保存到 {output_file_merged}")
    print(f"今天有 {max_value} 个涨停, 最高连板为 {max_lianban} 板, {max_lianban-1} 进 {max_lianban} 成功")



# 当日跌停获取
stock_zt_pool_dtgc_em_df = ak.stock_zt_pool_dtgc_em(date=date)

if not stock_zt_pool_dtgc_em_df.empty:

    stock_zt_pool_dtgc_em_df['成交额'] = (stock_zt_pool_dtgc_em_df['成交额'] / 100000000).round(2).astype(str) + '亿'
    stock_zt_pool_dtgc_em_df['流通市值'] = (stock_zt_pool_dtgc_em_df['流通市值'] / 100000000).round(2).astype(str) + '亿'
    stock_zt_pool_dtgc_em_df['总市值'] = (stock_zt_pool_dtgc_em_df['总市值'] / 100000000).round(2).astype(str) + '亿'
    stock_zt_pool_dtgc_em_df['封单资金'] = (stock_zt_pool_dtgc_em_df['封单资金'] / 100000000).round(2).astype(str) + '亿'
    stock_zt_pool_dtgc_em_df['涨跌幅'] = stock_zt_pool_dtgc_em_df['涨跌幅'].round(2).astype(str) + '%'
    stock_zt_pool_dtgc_em_df['换手率'] = stock_zt_pool_dtgc_em_df['换手率'].round(2).astype(str) + '%'
    stock_zt_pool_dtgc_em_df['最后封板时间'] = pd.to_datetime(stock_zt_pool_dtgc_em_df['最后封板时间'], format='%H%M%S').dt.time
    stock_zt_pool_dtgc_em_df['连续跌停'] = stock_zt_pool_dtgc_em_df['连续跌停'].astype(int)
    stock_zt_pool_dtgc_em_df.sort_values(by='连续跌停', ascending=False, inplace=True)

    # 改变涨跌幅颜色

    stock_zt_pool_dtgc_em_df.style
    stock_zt_pool_dtgc_em_df = stock_zt_pool_dtgc_em_df.style.map(color_negative_green_positive_red, subset=['涨跌幅'])

ws = book.create_sheet('跌停股池')
with pd.ExcelWriter(output_file_merged,mode='a',engine='openpyxl') as writer:
    stock_zt_pool_dtgc_em_df.to_excel(writer, sheet_name='跌停股池', index=False)
    print(f"跌停股池数据已成功保存到 {output_file_merged}")
    
